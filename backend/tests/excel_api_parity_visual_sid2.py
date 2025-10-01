#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel ↔ API Parity Visualizer (Auto-setup)
- If analysis exports are missing, auto-generates:
  * excel_scenario_parity_map_v1_2.json
  * parity_test_expected_values.csv
  from the provided Excel files (Blank + Filled).
- Then calls backend APIs and draws color-coded visual maps for Tax, TWC, Capex.

Usage:
  python excel_api_parity_visual_sid2.py --api-base http://127.0.0.1:8000 --scenario-id 2 \
         --blank-xlsm "Tender Model Blank.xlsm" --filled-xlsm "TEnder model filled.xlsm"

Requirements:
  pip install requests pandas matplotlib networkx pillow openpyxl
"""
import argparse, json, math, re
from pathlib import Path
from typing import Any, Dict, Optional, List
import requests
import pandas as pd
import matplotlib.pyplot as plt
import networkx as nx
from openpyxl import load_workbook

AREAS = ["Tax","TWC","Capex"]

def _is_num(x: Any) -> bool:
    try:
        float(x); return True
    except Exception:
        return False

def _norm_num(x: Any) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

def _compare_values(excel_val: Any, api_val: Any, eps: float = 1e-6) -> bool:
    if _is_num(excel_val) and _is_num(api_val):
        a, b = _norm_num(excel_val), _norm_num(api_val)
        return (a is not None and b is not None and abs(a-b) <= eps)
    return str(excel_val).strip() == str(api_val).strip()

def ensure_analysis_exports(blank_xlsm: Path, filled_xlsm: Path, out_dir: Path):
    """Create v1.2 parity map + expected values CSV if missing."""
    out_dir.mkdir(exist_ok=True, parents=True)
    v12 = out_dir / "excel_scenario_parity_map_v1_2.json"
    expected_csv = out_dir / "parity_test_expected_values.csv"
    if v12.exists() and expected_csv.exists():
        return v12, expected_csv

    # 1) Build explicit refs (heuristic) from Blank workbook
    wb_blank = load_workbook(filename=str(blank_xlsm), data_only=False, keep_vba=True)
    targets = {
        "Tax":   {"sheet_patterns": ["tax","vat","gst","kdv"],
                  "labels": ["vat","kdv","tax rate","gst","withholding","stopaj"]},
        "TWC":   {"sheet_patterns": ["twc","working","cash","capital","wc"],
                  "labels": ["dso","dpo","dio","inventory days","receivables days","payables days"]},
        "Capex": {"sheet_patterns": ["capex","asset","depr","depreciation","capital"],
                  "labels": ["capex","asset","spend","amount","useful life","depreciation","start month","start year"]},
    }
    def sheet_matches(t, pats): t=t.lower(); return any(p in t for p in pats)
    def scan_sheet(ws, labels, max_rows=200, max_cols=30):
        hints = []
        for r in range(1, min(ws.max_row, max_rows)+1):
            for c in range(1, min(ws.max_column, max_cols)+1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    s = v.strip().lower()
                    for lbl in labels:
                        if lbl in s and len(lbl) >= 3:
                            right = ws.cell(row=r, column=c+1) if c+1 <= max_cols else None
                            ref_cell = right if right is not None else ws.cell(row=r, column=c)
                            hints.append({
                                "label": lbl,
                                "found_text": v.strip(),
                                "label_cell": f"{ws.cell(row=r, column=c).coordinate}",
                                "value_cell": f"{ref_cell.coordinate}",
                                "sheet": ws.title,
                            })
                            if len(hints) >= 20:
                                return hints
        return hints

    explicit_refs = {k: [] for k in AREAS}
    for area, cfg in targets.items():
        for ws in wb_blank.worksheets:
            if sheet_matches(ws.title, cfg["sheet_patterns"]):
                explicit_refs[area].extend(scan_sheet(ws, cfg["labels"]))

    # 2) Build Parity v1.2 JSON
    module_map = {
        "Tax": {"backend":"scenario_tax.py","frontend":"TaxTab.tsx"},
        "TWC": {"backend":"twc.py","frontend":"TWCTab.tsx"},
        "Capex": {"backend":"scenario_capex.py","frontend":"CapexTable.tsx"},
    }
    parity = {}
    for area in AREAS:
        parity[area] = {
            "named_ranges": [],
            "backend_module": module_map[area]["backend"],
            "frontend_component": module_map[area]["frontend"],
            "manual_mapping": {
                "priority_named_ranges": [],
                "excel_sheet_hints": [],
                "excel_refs_examples": explicit_refs.get(area, [])
            }
        }
    v12.write_text(json.dumps(parity, indent=2, ensure_ascii=False), encoding="utf-8")

    # 3) Build expected CSV from Filled workbook
    wb_filled = load_workbook(filename=str(filled_xlsm), data_only=True, keep_vba=True)
    rows: List[Dict[str,Any]] = []
    for area in AREAS:
        for ref in parity[area]["manual_mapping"]["excel_refs_examples"]:
            try:
                ws = wb_filled[ref["sheet"]]
                val = ws[ref["value_cell"]].value
            except Exception:
                val = None
            rows.append({
                "Area": area,
                "Sheet": ref["sheet"],
                "Cell": ref["value_cell"],
                "Label": ref.get("found_text",""),
                "ExcelValue": val,
                "BackendEndpoint": parity[area]["backend_module"]
            })
    df = pd.DataFrame(rows)
    df.to_csv(expected_csv, index=False, encoding="utf-8")
    return v12, expected_csv

def fetch_api(area: str, api_base: str, scenario_id: int):
    session = requests.Session()
    paths = {
        "Tax":   [f"/api/scenarios/{scenario_id}/tax",   f"/scenarios/{scenario_id}/tax"],
        "TWC":   [f"/api/scenarios/{scenario_id}/twc",   f"/scenarios/{scenario_id}/twc"],
        "Capex": [f"/api/scenarios/{scenario_id}/capex", f"/scenarios/{scenario_id}/capex"],
    }.get(area, [])
    for path in paths:
        url = api_base.rstrip("/") + path
        try:
            r = session.get(url, timeout=10)
            if r.ok:
                return r.json()
        except Exception:
            continue
    return None

def extract_api_scalar(area: str, api_data: Any, label: str) -> Any:
    if api_data is None: return None
    try:
        if area == "Tax":
            for key in ("vat","vat_rate","tax_rate","gst","kdv"):
                if isinstance(api_data, dict) and key in api_data and _is_num(api_data[key]):
                    return api_data[key]
        if area == "TWC":
            for key in ("dso","dpo","dio","receivables_days","payables_days","inventory_days"):
                if isinstance(api_data, dict) and key in api_data and _is_num(api_data[key]):
                    return api_data[key]
        if area == "Capex":
            if isinstance(api_data, dict) and "items" in api_data and isinstance(api_data["items"], list):
                for item in api_data["items"]:
                    for key in ("amount","capex","value"):
                        if key in item and _is_num(item[key]):
                            return item[key]
        if isinstance(api_data, dict):
            for k, v in api_data.items():
                if _is_num(v): return v
    except Exception:
        return None
    return None

def draw_visuals(api_base: str, scenario_id: int, analysis_dir: Path, out_dir: Path):
    parity = json.loads((analysis_dir/"excel_scenario_parity_map_v1_2.json").read_text(encoding="utf-8"))
    df = pd.read_csv(analysis_dir/"parity_test_expected_values.csv")

    out_dir.mkdir(exist_ok=True, parents=True)
    for area in AREAS:
        sdf = df[df["Area"] == area].copy()
        if sdf.empty: 
            continue
        api_data = fetch_api(area, api_base, scenario_id)

        G = nx.DiGraph()
        api_node = f"{area} API\n/api/scenarios/{scenario_id}/{area.lower()}"
        G.add_node(api_node, type="backend")
        for _, row in sdf.iterrows():
            excel_node = f"{row['Sheet']}!{row['Cell']}\nExcel={row['ExcelValue']}"
            api_value = extract_api_scalar(area, api_data, str(row.get("Label","")))
            ok = _compare_values(row["ExcelValue"], api_value)
            color = "green" if ok else "red"
            G.add_node(excel_node, type="excel")
            G.add_edge(excel_node, api_node, color=color)

        pos = nx.spring_layout(G, k=0.7, iterations=60, seed=42)
        excel_nodes = [n for n,d in G.nodes(data=True) if d.get("type")=="excel"]
        backend_nodes = [n for n,d in G.nodes(data=True) if d.get("type")=="backend"]
        edge_colors = [d.get("color","gray") for _,_,d in G.edges(data=True)]

        plt.figure(figsize=(14,9))
        nx.draw_networkx_nodes(G, pos, nodelist=excel_nodes, node_size=600, label="Excel cells")
        nx.draw_networkx_nodes(G, pos, nodelist=backend_nodes, node_size=1400, label="Backend endpoint", node_shape="s")
        nx.draw_networkx_edges(G, pos, arrowstyle="->", arrowsize=12, edge_color=edge_colors)
        nx.draw_networkx_labels(G, pos, font_size=8)
        plt.title(f"Excel ↔ API Parity (Scenario {scenario_id}) — {area}\nGreen=match, Red=mismatch")
        plt.legend(scatterpoints=1, loc="upper left")
        plt.axis("off")
        out_img = out_dir / f"parity_visual_{area.lower()}_sid{scenario_id}.png"
        plt.savefig(out_img, dpi=150, bbox_inches="tight")
        plt.close()
        print(f"[OK] {out_img}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--api-base", default="http://127.0.0.1:8000")
    ap.add_argument("--scenario-id", type=int, default=2)
    ap.add_argument("--blank-xlsm", default="Tender Model Blank.xlsm")
    ap.add_argument("--filled-xlsm", default="TEnder model filled.xlsm")
    ap.add_argument("--analysis-dir", default="analysis_exports")
    ap.add_argument("--out-dir", default="parity_out")
    args = ap.parse_args()

    blank = Path(args.blank_xlsm).resolve()
    filled = Path(args.filled_xlsm).resolve()
    analysis_dir = Path(args.analysis_dir).resolve()
    out_dir = Path(args.out_dir).resolve()

    if not blank.exists():
        raise SystemExit(f"Blank workbook not found: {blank}")
    if not filled.exists():
        raise SystemExit(f"Filled workbook not found: {filled}")

    # Ensure inputs (or auto-generate them)
    ensure_analysis_exports(blank, filled, analysis_dir)

    # Draw visuals with live API comparisons
    draw_visuals(args.api_base, args.scenario_id, analysis_dir, out_dir)

if __name__ == "__main__":
    main()
